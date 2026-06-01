import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference


PROJECT_DIR = Path.cwd()

DATA_DIR = PROJECT_DIR / "data"
SQL_RESULTS_DIR = PROJECT_DIR / "sql_results"
REPORT_DIR = PROJECT_DIR / "report"

REPORT_DIR.mkdir(exist_ok=True)

OUTPUT_FILE = REPORT_DIR / "big_data_pet_announcements_report.xlsx"


CSV_SHEETS = [
    ("01_raw_data", DATA_DIR / "alertuj_ogloszenia_raw.csv"),
    ("02_clean_data", DATA_DIR / "alertuj_ogloszenia_clean.csv"),

    ("03_gold_animals", DATA_DIR / "gold_animal_counts.csv"),
    ("04_gold_regions", DATA_DIR / "gold_region_counts.csv"),
    ("05_gold_cities", DATA_DIR / "gold_city_counts.csv"),
    ("06_gold_rewards", DATA_DIR / "gold_reward_counts.csv"),
    ("07_gold_description", DATA_DIR / "gold_description_category_counts.csv"),
    ("08_gold_types", DATA_DIR / "gold_announcement_type_counts.csv"),
    ("09_gold_months", DATA_DIR / "gold_month_counts.csv"),
    ("10_gold_quality", DATA_DIR / "gold_data_quality_summary.csv"),

    ("11_sql_animals", SQL_RESULTS_DIR / "sql_animal_counts.csv"),
    ("12_sql_regions", SQL_RESULTS_DIR / "sql_region_counts.csv"),
    ("13_sql_cities", SQL_RESULTS_DIR / "sql_top_cities.csv"),
    ("14_sql_rewards", SQL_RESULTS_DIR / "sql_reward_counts.csv"),
    ("15_sql_months", SQL_RESULTS_DIR / "sql_month_counts.csv"),
    ("16_sql_description", SQL_RESULTS_DIR / "sql_description_quality.csv"),
    ("17_sql_types", SQL_RESULTS_DIR / "sql_announcement_type_counts.csv"),
    ("18_sql_quality", SQL_RESULTS_DIR / "sql_data_quality_summary.csv"),
]


def is_integer(value):
    value = str(value).strip()

    if value.startswith("-"):
        return value[1:].isdigit()

    return value.isdigit()


def is_float(value):
    value = str(value).strip()

    try:
        float(value)
        return "." in value
    except ValueError:
        return False


def convert_value(value):
    if value is None:
        return ""

    value = str(value).strip()

    if value == "":
        return ""

    if is_integer(value):
        return int(value)

    if is_float(value):
        return float(value)

    return value


def read_csv_file(file_path):
    rows = []

    if not file_path.exists():
        print("UWAGA: Nie znaleziono pliku:", file_path)
        return rows

    with open(file_path, "r", encoding="utf-8") as csvfile:
        reader = csv.reader(csvfile)

        for row in reader:
            converted_row = [convert_value(value) for value in row]
            rows.append(converted_row)

    return rows


def style_sheet(sheet):
    header_fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    thin_border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )

    if sheet.max_row >= 1:
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    for row in sheet.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            if isinstance(cell.value, int):
                cell.number_format = "0"

            if isinstance(cell.value, float):
                cell.number_format = "0.00"

    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)

        max_length = 0

        for cell in column_cells:
            value = cell.value

            if value is None:
                continue

            value_length = len(str(value))

            if value_length > max_length:
                max_length = value_length

        width = min(max(max_length + 2, 10), 45)
        sheet.column_dimensions[column_letter].width = width

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def add_bar_chart(sheet, title, x_col=1, y_col=2, position="D2"):
    if sheet.max_row < 2:
        return

    chart = BarChart()
    chart.title = title
    chart.y_axis.title = "Liczba"
    chart.x_axis.title = str(sheet.cell(row=1, column=x_col).value)

    data = Reference(
        sheet,
        min_col=y_col,
        min_row=1,
        max_row=sheet.max_row,
    )

    categories = Reference(
        sheet,
        min_col=x_col,
        min_row=2,
        max_row=sheet.max_row,
    )

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    chart.height = 8
    chart.width = 16

    sheet.add_chart(chart, position)


def add_pie_chart(sheet, title, x_col=1, y_col=2, position="D2"):
    if sheet.max_row < 2:
        return

    chart = PieChart()
    chart.title = title

    data = Reference(
        sheet,
        min_col=y_col,
        min_row=1,
        max_row=sheet.max_row,
    )

    labels = Reference(
        sheet,
        min_col=x_col,
        min_row=2,
        max_row=sheet.max_row,
    )

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)

    chart.height = 8
    chart.width = 12

    sheet.add_chart(chart, position)


def create_workbook():
    workbook = Workbook()

    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for sheet_name, csv_file in CSV_SHEETS:
        rows = read_csv_file(csv_file)

        sheet = workbook.create_sheet(title=sheet_name)

        if not rows:
            sheet.append(["Brak danych"])
            continue

        for row in rows:
            sheet.append(row)

        style_sheet(sheet)

    return workbook


def add_charts(workbook):
    # Wykresy dla warstwy Gold
    if "03_gold_animals" in workbook.sheetnames:
        add_pie_chart(
            workbook["03_gold_animals"],
            "Gold: ogłoszenia według gatunku",
            position="D2",
        )

    if "04_gold_regions" in workbook.sheetnames:
        add_bar_chart(
            workbook["04_gold_regions"],
            "Gold: top województw",
            position="D2",
        )

    if "06_gold_rewards" in workbook.sheetnames:
        add_pie_chart(
            workbook["06_gold_rewards"],
            "Gold: ogłoszenia z nagrodą i bez nagrody",
            position="D2",
        )

    if "07_gold_description" in workbook.sheetnames:
        add_bar_chart(
            workbook["07_gold_description"],
            "Gold: długość opisów",
            position="D2",
        )

    if "08_gold_types" in workbook.sheetnames:
        add_pie_chart(
            workbook["08_gold_types"],
            "Gold: typy ogłoszeń",
            position="D2",
        )

    # Wykresy dla warstwy SQL
    if "11_sql_animals" in workbook.sheetnames:
        add_pie_chart(
            workbook["11_sql_animals"],
            "SQL: ogłoszenia według gatunku",
            position="D2",
        )

    if "12_sql_regions" in workbook.sheetnames:
        add_bar_chart(
            workbook["12_sql_regions"],
            "SQL: top województw",
            position="D2",
        )

    if "14_sql_rewards" in workbook.sheetnames:
        add_pie_chart(
            workbook["14_sql_rewards"],
            "SQL: ogłoszenia z nagrodą i bez nagrody",
            position="D2",
        )

    if "16_sql_description" in workbook.sheetnames:
        add_bar_chart(
            workbook["16_sql_description"],
            "SQL: jakość opisów",
            position="E2",
        )

    if "17_sql_types" in workbook.sheetnames:
        add_pie_chart(
            workbook["17_sql_types"],
            "SQL: typy ogłoszeń",
            position="D2",
        )


def print_summary(workbook):
    print("Utworzono plik Excel:", OUTPUT_FILE)
    print("Arkusze w pliku:")

    for sheet_name in workbook.sheetnames:
        print("-", sheet_name)


def main():
    workbook = create_workbook()
    add_charts(workbook)

    workbook.save(OUTPUT_FILE)

    print_summary(workbook)


if __name__ == "__main__":
    main()
